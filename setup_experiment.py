#!/usr/bin/env python3
"""
Setup a new clustering experiment for the UI.

Usage:
    python3 setup_experiment.py \
        --name HT0601 \
        --label "HT 2026-06-01" \
        --visual_csv path/to/cluster_result_visual.csv \
        --audio_csv  path/to/cluster_result_audio.csv \
        --lasso_audio  path/to/lasso_v20_audio-only_result.xlsx \
        --lasso_visual path/to/lasso_v20_visual-only_result.xlsx \
        --group_name_visual path/to/group_name_visual.xlsx \
        --wordcloud_dir     path/to/wordcloud/ \
        [--json_desc ./data/json_description]   # default
        [--output_base ./data/experiments]       # default
"""

import argparse, json, glob, os, shutil, re
import numpy as np
import pandas as pd
from collections import Counter


# ── Helpers ───────────────────────────────────────────────────────────────────

def last_seg(s):
    return str(s).split("-")[-1].replace(".mp4", "").strip()


def agg_numeric(vals):
    arr = [v for v in vals if isinstance(v, (int, float)) and v is not None]
    if not arr:
        return None
    return {
        "min":  float(np.min(arr)),
        "max":  float(np.max(arr)),
        "mean": round(float(np.mean(arr)), 2),
        "avg":  round(float(np.mean(arr)), 2),
        "std":  round(float(np.std(arr)), 2),
    }


def agg_categorical(vals):
    counts = Counter()
    for v in vals:
        if v is None:
            continue
        if isinstance(v, list):
            for item in v:
                if item:
                    counts[str(item)] += 1
        else:
            counts[str(v)] += 1
    return dict(counts.most_common(20))


def is_scale_int(vals):
    nums = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
    return len(nums) > len(vals) * 0.5


def aggregate_section(videos_data, section_path):
    parts = section_path.split(".")
    samples = []
    for vd in videos_data:
        node = vd
        for p in parts:
            node = node.get(p) if isinstance(node, dict) else None
        if isinstance(node, dict):
            samples.append(node)
    if not samples:
        return {}
    all_keys = {k for s in samples for k in s}
    result = {}
    for key in all_keys:
        vals = [s[key] for s in samples if key in s]
        # skip nested dicts — they are handled as their own sub-section
        non_dict = [v for v in vals if not isinstance(v, dict)]
        if not non_dict:
            continue
        result[key] = agg_numeric(non_dict) if is_scale_int(non_dict) else agg_categorical(non_dict)
    return {k: v for k, v in result.items() if v is not None}


def discover_sections(media_map):
    """Recursively walk video JSONs and collect every dict-node path as a section.

    Works with any JSON schema: auto-discovers all top-level dict keys so it
    handles both v07 (low_inference_observations) and v20 (low_inference /
    high_inference) without any hardcoded paths.
    Section names strip the top-level container key to stay concise.
    """
    def get_node(d, path_parts):
        for p in path_parts:
            d = d.get(p) if isinstance(d, dict) else None
        return d

    def walk(path_parts):
        sections = {}
        nodes = [get_node(v, path_parts) for v in media_map.values()]
        nodes = [n for n in nodes if isinstance(n, dict)]
        if not nodes:
            return sections

        has_scalar = any(
            not isinstance(n.get(k), dict)
            for n in nodes for k in n
            if n.get(k) is not None
        )
        if has_scalar:
            full_path = ".".join(path_parts)
            # Strip the outermost container key to keep names concise
            rel_parts = path_parts[1:]
            section_name = "_".join(rel_parts) if rel_parts else path_parts[0]
            sections[section_name] = full_path

        dict_keys = sorted({k for n in nodes for k, v in n.items() if isinstance(v, dict)})
        for subkey in dict_keys:
            sections.update(walk(path_parts + [subkey]))

        return sections

    # Auto-detect all top-level dict-valued keys across videos
    sample = next(iter(media_map.values()), {})
    top_keys = sorted(k for k, v in sample.items() if isinstance(v, dict))
    all_sections = {}
    for key in top_keys:
        all_sections.update(walk([key]))
    return all_sections


# ── Core converters ───────────────────────────────────────────────────────────

def build_media_map(json_desc_dir):
    media_map = {}
    for f in glob.glob(os.path.join(json_desc_dir, "*.json")):
        with open(f, encoding="utf-8") as fp:
            d = json.load(fp)
        for key, val in d.items():
            if isinstance(val, dict):
                lid = last_seg(val.get("videoName", key))
                media_map[lid] = val
    return media_map


def normalize_cluster_csv(csv_path, type_name):
    """Load CSV and normalize column names for Python + JS compatibility.

    Handles two formats:
      old — file, {type}_kmeans_lables, id
      new — file_id, label_id, label, [text]
    Returns (normalized_df, label_col).
    """
    df = pd.read_csv(csv_path)
    label_col = f"{type_name}_kmeans_lables"

    if "file" not in df.columns and "file_id" in df.columns:
        df = df.rename(columns={"file_id": "file"})

    if label_col not in df.columns:
        if "label_id" in df.columns:
            df = df.rename(columns={"label_id": label_col})
        elif "label" in df.columns:
            df = df.rename(columns={"label": label_col})

    # id column = ig_id (first segment) used by cluster.js for creator lookup
    if "id" not in df.columns and "file" in df.columns:
        df["id"] = df["file"].apply(lambda x: str(x).split("-")[0])

    return df, label_col


def generate_cluster_compare(df, label_col, media_map, sections):
    df["_lid"] = df["file"].apply(last_seg)
    output = {}
    for label in sorted(df[label_col].unique()):
        group_df = df[df[label_col] == label]
        videos = [media_map[lid] for lid in group_df["_lid"] if lid in media_map]
        n_total, n_matched = len(group_df), len(videos)
        print(f"    cluster {label}: {n_total} videos, {n_matched} matched")
        stats = {k: v for k, v in
                 ((k, aggregate_section(videos, p)) for k, p in sections.items()) if v}
        output[f"cluster_{label}"] = {
            "metadata": {"cluster_label": str(label),
                         "video_count": n_total, "matched_count": n_matched},
            "statistics": stats,
        }
    return output


def convert_lasso_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        result[sheet] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = dict(zip(headers, row))
            cid = str(d["cluster"])
            result[sheet][cid] = {
                "pos": [d.get(f"pos_{i}") for i in range(1, 6) if d.get(f"pos_{i}")],
                "neg": [d.get(f"neg_{i}") for i in range(1, 6) if d.get(f"neg_{i}")],
            }
    return result


def convert_group_name_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        label = str(d["group_label"])
        result[label] = {
            "group_name_GPT":    d.get("group_name-GPT") or "",
            "group_des_GPT":     d.get("group_des-GPT") or "",
            "group_name_Gemini": d.get("group_name-Gemini") or "",
            "group_des_Gemini":  d.get("group_des-Gemini") or "",
        }
    return result


def copy_wordcloud(wc_dir, dest_dir):
    """
    Detect keybert / tfidf PNGs by regex, copy to dest_dir as
    keybert_cluster_{n}.png / tfidf_cluster_{n}.png
    """
    filter_word = ''
    if 'audio' in dest_dir:
        filter_word = 'audio'
    elif 'visual' in dest_dir:
        filter_word = 'visual'
    os.makedirs(dest_dir, exist_ok=True)
    copied = 0
    for f in glob.glob(os.path.join(wc_dir, "*.png")):
        fname = os.path.basename(f).lower()
        for method in ("keybert", "tfidf"):
            if method in fname and filter_word in fname:
                m = re.search(r"cluster[_\-]?(\d+)", fname)
                if m:
                    dst = os.path.join(dest_dir, f"{method}_cluster_{m.group(1)}.png")
                    shutil.copy2(f, dst)
                    copied += 1
    return copied


def update_manifest(manifest_path, exp_id, label, types):
    if os.path.exists(manifest_path):
        with open(manifest_path, encoding="utf-8") as fp:
            manifest = json.load(fp)
    else:
        manifest = {"experiments": [], "default": exp_id}

    existing = next((e for e in manifest["experiments"] if e["id"] == exp_id), None)
    if existing:
        existing["label"] = label
        existing["types"] = types
    else:
        manifest["experiments"].append({"id": exp_id, "label": label, "types": types})

    if not manifest.get("default"):
        manifest["default"] = exp_id

    with open(manifest_path, "w", encoding="utf-8") as fp:
        json.dump(manifest, fp, ensure_ascii=False, indent=2)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Setup clustering experiment for UI")
    parser.add_argument("--name",   required=True, help="Experiment ID, e.g. HT0601")
    parser.add_argument("--label",  default=None,  help='Display label, e.g. "HT 2026-06-01"')
    parser.add_argument("--visual_csv",        default=None)
    parser.add_argument("--audio_csv",         default=None)
    parser.add_argument("--lasso_audio",        default=None, help="lasso result xlsx for audio")
    parser.add_argument("--lasso_visual",       default=None, help="lasso result xlsx for visual")
    parser.add_argument("--group_name_visual", default=None)
    parser.add_argument("--group_name_audio",  default=None)
    parser.add_argument("--wordcloud_dir",     default=None)
    parser.add_argument("--json_desc",   default="./data/json_description")
    parser.add_argument("--output_base", default="./data/experiments")
    args = parser.parse_args()

    exp_id   = args.name
    label    = args.label or exp_id
    exp_dir  = os.path.join(args.output_base, exp_id)
    os.makedirs(exp_dir, exist_ok=True)

    # Build media map
    print(f"Loading json_description from {args.json_desc} ...")
    media_map = build_media_map(args.json_desc)
    print(f"  {len(media_map)} videos loaded")
    sections = discover_sections(media_map)
    print(f"  {len(sections)} sections discovered: {list(sections.keys())}")

    # Convert lasso xlsx → shared lasso_features.json
    lasso_data = {}
    for lasso_path in filter(None, [args.lasso_audio, args.lasso_visual]):
        print(f"\nConverting lasso xlsx: {lasso_path}")
        lasso_data.update(convert_lasso_xlsx(lasso_path))
    if lasso_data:
        out = os.path.join(exp_dir, "lasso_features.json")
        with open(out, "w", encoding="utf-8") as fp:
            json.dump(lasso_data, fp, ensure_ascii=False, indent=2)
        print(f"  → {out}")

    # Process each type
    type_cfgs = []
    if args.visual_csv:
        type_cfgs.append(("visual", args.visual_csv, args.group_name_visual))
    if args.audio_csv:
        type_cfgs.append(("audio",  args.audio_csv,  args.group_name_audio))

    available_types = []
    for type_name, csv_path, group_name_path in type_cfgs:
        print(f"\nProcessing {type_name} ...")
        type_dir = os.path.join(exp_dir, type_name)
        os.makedirs(type_dir, exist_ok=True)

        # Normalize and save cluster CSV
        df_csv, label_col = normalize_cluster_csv(csv_path, type_name)
        dst = os.path.join(type_dir, "cluster_result.csv")
        df_csv.to_csv(dst, index=False)
        print(f"  Saved cluster CSV → {dst} (label_col={label_col})")

        # Generate cluster_compare.json
        print(f"  Generating cluster_compare.json ...")
        compare = generate_cluster_compare(df_csv, label_col, media_map, sections)
        out = os.path.join(type_dir, "cluster_compare.json")
        with open(out, "w", encoding="utf-8") as fp:
            json.dump(compare, fp, ensure_ascii=False, indent=2)
        print(f"  → {out}")

        # Convert group_name xlsx
        if group_name_path:
            group_names = convert_group_name_xlsx(group_name_path)
            out = os.path.join(type_dir, "group_names.json")
            with open(out, "w", encoding="utf-8") as fp:
                json.dump(group_names, fp, ensure_ascii=False, indent=2)
            print(f"  → group_names: {out}")

        # Copy wordcloud images
        if args.wordcloud_dir:
            wc_dest = os.path.join(type_dir, "wordcloud")
            n = copy_wordcloud(args.wordcloud_dir, wc_dest)
            print(f"  → wordcloud: {n} images → {wc_dest}")

        available_types.append(type_name)

    # Update manifest
    manifest_path = os.path.join(args.output_base, "manifest.json")
    update_manifest(manifest_path, exp_id, label, available_types)
    print(f"\n✓ Manifest updated: {manifest_path}")
    print(f"✓ Experiment '{exp_id}' ready at {exp_dir}")
    print(f"  Types: {available_types}")


if __name__ == "__main__":
    main()
